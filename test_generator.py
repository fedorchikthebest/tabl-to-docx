from docxtpl import DocxTemplate
from openpyxl import load_workbook
import docx
from docx2pdf import convert
import os
from sendmail import send_mail


def to_kavichki(text:str) -> str:
    text = str(text)
    while '"' in text:
        text = text.replace('"', "«", 1)
        text = text[::-1].replace('"', "»", 1)[::-1]
    return text


classes = [
    'работник, назначенный в качестве лица, ответственного за обеспечение транспортной безопасности в субъекте транспортной инфраструктуры',
    'работник, назначенный в качестве лица, ответственного за обеспечение транспортной безопасности на объекте транспортной инфраструктуры и (или) транспортном средстве, и персонала специализированных организаций',
    'работник субъекта транспортной инфраструктуры, подразделения транспортной безопасности, руководящий выполнением работы, непосредственно связанной с обеспечением транспортной безопасности объекта транспортной инфраструктуры и (или) транспортного средства',
    'работник, включенный в состав группы быстрого реагирования',
    'работник, осуществляющий досмотр, дополнительный досмотр и повторный досмотр в целях обеспечения транспортной безопасности',
    'работник, осуществляющий наблюдение и (или) собеседование в целях обеспечения транспортной безопасности',
    'работник, управляющий техническими средствами обеспечения транспортной безопасности',
    'иной работник субъекта транспортной инфраструктуры, подразделения транспортной безопасности, выполняющий работы, непосредственно связанные с обеспечением транспортной безопасности объекта транспортной инфраструктуры и (или) транспортного средства'
]


def render_shablons(csv_path, save_path):
    wb_obj = load_workbook(csv_path)
    sheet_obj = wb_obj.active
    
    row_num = 2
    
    while sheet_obj.cell(row = row_num, column = 1).value is not None:
        doc = DocxTemplate("shablon.docx")
        doc_reshenie = DocxTemplate("attest_shablon.docx")
        contexts = [sheet_obj.cell(row_num, i).value for i in range(1,10)]
        context = {'name' : to_kavichki(contexts[0]),
                'born_date' : contexts[1],
                'a_class' : classes[int(contexts[2]) - 1],
                'attes_num' : to_kavichki(contexts[3]),
                'start_date' : contexts[4],
                'stop_date' : contexts[5],
                'reshenie_date': contexts[6],
                'company': to_kavichki(contexts[7]),
                'email': to_kavichki(contexts[8])}
        
        reshenie_name = f"{save_path}\\reshenie{context['attes_num']}_{row_num}.pdf"
        docx_name = f"{save_path}\\{context['attes_num']}_{row_num}"

        doc.render(context)
        doc.save(f"{docx_name}.docx")
        
        doc = docx.Document(f"{docx_name}.docx")
        doc.save(f"{docx_name}.doc")
        os.remove(f"{docx_name}.docx")
        
        doc_reshenie.render(context)
        doc_reshenie.save(f"{reshenie_name}.docx")
        convert(f"{reshenie_name}.docx", f"{reshenie_name}.pdf")
        os.remove(f"{reshenie_name}.docx")
        try:
            send_mail(context['email'], "Решение", f"{reshenie_name}.pdf")
        except Exception:
            print('Не удалось отправить письмо', context['email'])
        
        
        row_num += 1


render_shablons("./data.xlsx", "./result")
