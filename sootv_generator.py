from docxtpl import DocxTemplate
from openpyxl import load_workbook
from docx2pdf import convert
import os
from sendmail import send_mail


def to_kavichki(text:str) -> str:
    text = str(text)
    while '"' in text:
        text = text.replace('"', "«", 1)
        text = text[::-1].replace('"', "»", 1)[::-1]
    return text


classes = [
    'работник, назначенный в качестве лица, ответственного за обеспечение транспортной безопасности в субъекте транспортной инфраструктуры',
    'работник, назначенный в качестве лица, ответственного за обеспечение транспортной безопасности на объекте транспортной инфраструктуры и (или) транспортном средстве, и персонала специализированных организаций',
    'работник субъекта транспортной инфраструктуры, подразделения транспортной безопасности, руководящий выполнением работы, непосредственно связанной с обеспечением транспортной безопасности объекта транспортной инфраструктуры и (или) транспортного средства',
    'работник, включенный в состав группы быстрого реагирования',
    'работник, осуществляющий досмотр, дополнительный досмотр и повторный досмотр в целях обеспечения транспортной безопасности',
    'работник, осуществляющий наблюдение и (или) собеседование в целях обеспечения транспортной безопасности',
    'работник, управляющий техническими средствами обеспечения транспортной безопасности',
    'иной работник субъекта транспортной инфраструктуры, подразделения транспортной безопасности, выполняющий работы, непосредственно связанные с обеспечением транспортной безопасности объекта транспортной инфраструктуры и (или) транспортного средства'
]


def render_shablons(csv_path, save_path):
    wb_obj = load_workbook(csv_path)
    sheet_obj = wb_obj.active
    
    doc = DocxTemplate("sootv_shablon.docx")
    row_num = 2
    
    while sheet_obj.cell(row = row_num, column = 1).value is not None:
        contexts = [sheet_obj.cell(row_num, i).value for i in range(1,10)]
        context = {'company' : to_kavichki(contexts[0]),
                'name' : contexts[1],
                'num' :  to_kavichki(contexts[2]),
                'a_class' : classes[int(contexts[3]) - 1],
                "email": contexts[4]}

        doc.render(context)
        doc.save(f"{save_path}\\{context['num']}_{row_num}.docx")
        
        convert(f"{save_path}\\{context['num']}_{row_num}.docx", f"{save_path}\\{context['num']}_{row_num}.pdf")
        os.remove(f"{save_path}\\{context['num']}_{row_num}.docx")
        
        try:
            send_mail(context['email'], "Решение соответствие", f"{save_path}\\{context['num']}_{row_num}.pdf")    
        except Exception:
            print('Не удалось отправить письмо', context['email'])
        
        row_num += 1


render_shablons("./data_sootv.xlsx", "./sootv")
